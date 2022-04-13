import json
import openpyxl
import datetime
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Side, Border

class cXLWorkBook:
    def __init__(self, filepath, proj_name, data_json):
        self.filepath = filepath
        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.title = proj_name + " 시험정보"

        worksheet.append(["시험정보"])
        worksheet.append([])
        worksheet.append(["1. 데이터셋 경로", data_json['target_directory']])
        worksheet.column_dimensions['A'].width = 20

        time_now = datetime.datetime.now()
        date_string = time_now.strftime('%Y-%m-%d %H:%M')

        worksheet.append(["2. 시험일시", date_string])

        worksheet['A1'].font = Font(bold=True)
        worksheet['A3'].font = Font(bold=True)
        worksheet['A4'].font = Font(bold=True)

        worksheet_statistics = wb.create_sheet()
        worksheet_statistics.title = proj_name + " 통계적다양성"

        worksheet_syntax = wb.create_sheet()
        worksheet_syntax.title = proj_name + " 구문정확성"

        schema_property_list = []
        schema_property_name_list = []
        schema_property_type_list = []
        schema_property_path_list = []

        grayFill = PatternFill(start_color='cccccc',
            end_color='cccccc',
            fill_type='solid')

        col_index = 2
        max_row_index = 1
        STATISTICS_VALUE_START_ROW = 4
        worksheet_statistics.cell(1, 1).value = "항목명"
        worksheet_statistics.cell(2, 1).value = "타입"
        worksheet_statistics.cell(3, 1).value = "경로"
        worksheet_statistics.cell(4, 1).value = "시험결과"
        worksheet_statistics.cell(4, 1).alignment = Alignment(vertical='center')
        worksheet_statistics['A1'].font = Font(bold=True)
        worksheet_statistics['A2'].font = Font(bold=True)
        worksheet_statistics['A3'].font = Font(bold=True)
        worksheet_statistics['A4'].font = Font(bold=True)
        worksheet_statistics['A1'].fill = grayFill

        worksheet_syntax.cell(1, 1).value = "파일 명"
        worksheet_syntax.cell(1, 2).value = "결과"
        worksheet_syntax.cell(1, 3).value = "구문 통계 오류"
        worksheet_syntax['A1'].font = Font(bold=True)
        worksheet_syntax['B1'].font = Font(bold=True)
        worksheet_syntax['C1'].font = Font(bold=True)
        worksheet_syntax['A1'].fill = grayFill
        worksheet_syntax['B1'].fill = grayFill
        worksheet_syntax['C1'].fill = grayFill

        for schema_property_key, schema_property in data_json['schema_property'].items():
            schema_property_list.append(schema_property_key)
            #schema_property_name_list.append(schema_property_key.split(',')[-1])
            #schema_property_type_list.append(schema_property['type'])
            #schema_property_path_list.append(schema_property_key)

            worksheet_statistics.cell(1, col_index).value = schema_property_key.split('.')[-1]
            worksheet_statistics.cell(1, col_index).fill = grayFill

            worksheet_statistics.cell(2, col_index).value = schema_property['type']
            worksheet_statistics.cell(3, col_index).value = schema_property_key
            worksheet_statistics[chr(65+col_index-1)+str(1)].font = Font(bold=True)
            worksheet_statistics[chr(65+col_index-1)+str(2)].font = Font(bold=True)
            worksheet_statistics[chr(65+col_index-1)+str(3)].font = Font(bold=True)

            worksheet_statistics.merge_cells(start_row=1, start_column=col_index, end_row=1, end_column=col_index+1)
            worksheet_statistics.merge_cells(start_row=2, start_column=col_index, end_row=2, end_column=col_index+1)
            worksheet_statistics.merge_cells(start_row=3, start_column=col_index, end_row=3, end_column=col_index+1)

            worksheet_statistics.column_dimensions[chr(65+col_index-1)].width = 30

            if 'value_list' not in schema_property:
                valid_value_list = None
            else:
                valid_value_list = schema_property['value_list'].split(",")

            row_index = STATISTICS_VALUE_START_ROW + 1
            total_count = 0
            #print("schema_property['value_count']")
            #print(json.dumps(schema_property['value_count'], indent=4, ensure_ascii=False))
            #keys1 = schema_property['value_count'].keys()
            #print(schema_property['value_count'].keys())
            #for v in keys1:
            #    print(type(v))
            #    print(v)
            #keys1 = sorted(keys1)
            for value_key in sorted(schema_property['value_count'].keys()):
                #print(value_key)
                value_count = schema_property['value_count'][value_key]
                if valid_value_list != None and value_key not in valid_value_list:
                    continue
                worksheet_statistics.cell(row_index, col_index + 0).value = value_key
                worksheet_statistics.cell(row_index, col_index + 1).value = value_count
                worksheet_statistics[chr(65+col_index-1)+str(row_index)].font = Font(bold=True)

                total_count = total_count + value_count
                row_index = row_index + 1
            redFill = PatternFill(start_color='FF0000',
                end_color='FF0000',
                fill_type='solid')
            for value_key, value_count in schema_property['value_count'].items():
                if valid_value_list != None and value_key in valid_value_list:
                    continue
                if valid_value_list == None:
                    continue
                worksheet_statistics.cell(row_index, col_index + 0).value = value_key
                worksheet_statistics.cell(row_index, col_index + 1).value = value_count
                worksheet_statistics.cell(row_index, col_index + 0).fill = redFill
                worksheet_statistics.cell(row_index, col_index + 1).fill = redFill
                worksheet_statistics[chr(65+col_index-1)+str(row_index)].font = Font(bold=True)

                total_count = total_count + value_count

                row_index = row_index + 1

            if max_row_index < row_index:
                max_row_index = row_index

            # total count
            worksheet_statistics.cell(STATISTICS_VALUE_START_ROW, col_index + 0).value = "total"
            worksheet_statistics.cell(STATISTICS_VALUE_START_ROW, col_index + 1).value = total_count
            worksheet_statistics[chr(65+col_index-1)+str(STATISTICS_VALUE_START_ROW)].font = Font(bold=True)

            col_index = col_index + 2

        row_index = 2
        for file_path, file_validation_result in data_json['file_list'].items():
            worksheet_syntax.cell(row_index, 1).value = file_path
            worksheet_syntax.cell(row_index, 2).value = \
                "PASS" if file_validation_result['valid_schema']==True else "FAIL"
            worksheet_syntax.cell(row_index, 3).value = \
                "" if file_validation_result['valid_schema']==True else file_validation_result["schema_validation_msg"]

            if file_validation_result['valid_schema']!=True:
                worksheet_syntax.cell(row_index, 2).fill = redFill
                worksheet_syntax.cell(row_index, 3).fill = redFill

            row_index = row_index + 1

        worksheet_statistics.merge_cells(start_row=4, start_column=1, end_row=max_row_index-1, end_column=1)
        worksheet_syntax = cXLWorkBook.AutoFitColumnSize(worksheet_syntax, None, 10)

        wb.save(filepath)
        wb.close()

    def AutoFitColumnSize(worksheet, columns=None, margin=2):
        for i, column_cells in enumerate(worksheet.columns):
            is_ok = False
            if columns == None:
                is_ok = True
            elif isinstance(columns, list) and i in columns:
                is_ok = True

            if is_ok:
                length = max(len(str(cell.value)) for cell in column_cells)
                worksheet.column_dimensions[column_cells[0]
                                            .column_letter].width = length + margin

        return worksheet

    def __del__(self):
        pass